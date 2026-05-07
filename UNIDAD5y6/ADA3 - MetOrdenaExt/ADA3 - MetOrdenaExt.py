"""
Intercalación de Archivos
--------------------------
Carga 2 archivos (.txt o .json) con números desordenados,
los une y los ordena usando Insertion Sort (Intercalación).
Guarda el resultado en un nuevo archivo.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import os
import random

# ── Colores ──────────────────────────────────────────────
BG        = "#0f1117"
CARD      = "#1a1d27"
PANEL     = "#141720"
BLUE      = "#4f8ef7"
GREEN     = "#3ecf8e"
GOLD      = "#f5c842"
PINK      = "#f75f8e"
TEXT      = "#e8eaf0"
MUTED     = "#7b8094"
BAR_DEF   = "#4f8ef7"
BAR_KEY   = "#f5c842"
BAR_CMP   = "#f75f8e"
BAR_DONE  = "#3ecf8e"


# ── Insertion Sort con pasos ──────────────────────────────
def insertion_sort_steps(arr):
    a = arr[:]
    steps = []
    n = len(a)
    for i in range(1, n):
        key = a[i]
        j = i - 1
        steps.append(("key", i, a[:], f"Clave tomada: a[{i}] = {key}"))
        while j >= 0 and a[j] > key:
            steps.append(("cmp", j, a[:], f"a[{j}]={a[j]} > {key} → desplazar"))
            a[j + 1] = a[j]
            j -= 1
            steps.append(("move", j + 1, a[:], f"Elemento desplazado a posición {j+2}"))
        a[j + 1] = key
        steps.append(("place", j + 1, a[:], f"Insertado {key} en posición {j+1}"))
    steps.append(("done", -1, a[:], "¡Ordenamiento completado!"))
    return steps


# ── Lectura de archivos ───────────────────────────────────
# Guarda info de hojas leídas por archivo xlsx
_xlsx_info = {}

def leer_archivo(path):
    """Lee un .txt, .json, .csv o .xlsx y devuelve lista de números."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".json":
        with open(path, "r", encoding="utf-8") as f:
            datos = json.load(f)
        if isinstance(datos, list):
            return [float(x) for x in datos]
        raise ValueError("El JSON debe contener una lista de números.\nEjemplo: [34, 7, 23, 90, 1]")

    elif ext == ".csv":
        import csv
        numeros = []
        with open(path, newline="", encoding="utf-8") as f:
            # Detectar si tiene encabezado
            sample = f.read(2048); f.seek(0)
            has_header = csv.Sniffer().has_header(sample)
            reader = csv.reader(f)
            if has_header:
                next(reader)   # saltar encabezado
            for row in reader:
                for cell in row:
                    cell = cell.strip()
                    if cell:
                        try:
                            numeros.append(float(cell))
                        except ValueError:
                            pass   # ignorar texto no numérico
        if not numeros:
            raise ValueError("No se encontraron números en el archivo CSV.")
        return numeros

    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Instala openpyxl: pip install openpyxl")
        
        # Quitamos data_only=True si queremos ver fórmulas, 
        # pero lo dejamos para obtener los valores calculados
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        numeros = []
        hojas_leidas = []
        tipos_conteo = {"Números": 0, "Texto": 0, "Booleanos": 0, "Otros": 0}
        
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            encontrados_en_hoja = 0
            for row in ws.iter_rows(): # Quitamos values_only para acceder al objeto cell
                for cell in row:
                    # --- Identificación del tipo de dato ---
                    t = cell.data_type
                    if t == 'n': tipos_conteo["Números"] += 1
                    elif t == 's': tipos_conteo["Texto"] += 1
                    elif t == 'b': tipos_conteo["Booleanos"] += 1
                    else: tipos_conteo["Otros"] += 1
                    
                    val = cell.value
                    if val is None or isinstance(val, (bool, str)):
                        continue
                    if isinstance(val, (int, float)):
                        numeros.append(float(val))
                        encontrados_en_hoja += 1
            
            if encontrados_en_hoja > 0:
                hojas_leidas.append(f"{nombre_hoja} ({encontrados_en_hoja})")
        
        wb.close()
        if not numeros:
            raise ValueError("No se encontraron números en el archivo XLSX.")
        
        # Guardar info extendida
        resumen_tipos = f"N: {tipos_conteo['Números']}, T: {tipos_conteo['Texto']}, B: {tipos_conteo['Booleanos']}"
        _xlsx_info[path] = f"{', '.join(hojas_leidas)} | Tipos -> {resumen_tipos}"
        
        return numeros

    else:  # .txt
        import re
        with open(path, "r", encoding="utf-8") as f:
            contenido = f.read().strip()
        tokens = re.split(r"[,\s\n]+", contenido)
        numeros = []
        for t in tokens:
            if t:
                try:
                    numeros.append(float(t))
                except ValueError:
                    pass
        if not numeros:
            raise ValueError("No se encontraron números en el archivo TXT.")
        return numeros


def guardar_archivo(path, datos):
    ext = os.path.splitext(path)[1].lower()
    with open(path, "w", encoding="utf-8") as f:
        if ext == ".json":
            json.dump([int(x) if x == int(x) else x for x in datos], f, indent=2)
        else:
            f.write(", ".join(str(int(x) if x == int(x) else x) for x in datos))


# ── Canvas de barras ──────────────────────────────────────
class BarCanvas(tk.Canvas):
    def __init__(self, master, **kw):
        super().__init__(master, bg=CARD, highlightthickness=0, **kw)
        self.data = []
        self.hl   = {}

    def set_data(self, data, hl=None):
        self.data = data
        self.hl   = hl or {}
        self.redraw()

    def redraw(self):
        self.delete("all")
        if not self.data:
            return
        w = self.winfo_width()  or 700
        h = self.winfo_height() or 180
        n = len(self.data)
        max_v  = max(self.data) if self.data else 1
        pad    = 10
        bar_w  = max(2, (w - 2*pad) / n - 1)

        for i, val in enumerate(self.data):
            x0 = pad + i * (bar_w + 1)
            x1 = x0 + bar_w
            bh = max(4, int((val / max_v) * (h - 28)))
            y0 = h - bh - 10
            y1 = h - 10
            color = self.hl.get(i, BAR_DEF)
            self.create_rectangle(x0+2, y0+2, x1+2, y1+2, fill="#0a0c12", outline="")
            self.create_rectangle(x0, y0, x1, y1, fill=color, outline="")
            if bar_w >= 20:
                self.create_text((x0+x1)//2, y0-7, text=str(int(val)),
                                  fill=MUTED, font=("Consolas", 7))


# ── Aplicación principal ──────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Intercalación de Archivos — Insertion Sort")
        self.configure(bg=BG)
        self.geometry("860x680")
        self.resizable(True, True)

        self.file1  = tk.StringVar(value="Sin archivo")
        self.file2  = tk.StringVar(value="Sin archivo")
        self.data1  = []
        self.data2  = []
        self.merged = []
        self.steps  = []
        self.idx    = 0
        self.running = False
        self._after  = None
        self.speed   = 80

        self._build()

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Horizontal.TProgressbar",
                         background=BLUE, troughcolor=PANEL,
                         borderwidth=0, thickness=5)

    # ── Construcción de la UI ─────────────────────────────
    def _build(self):
        # Título
        tk.Label(self, text="Intercalación de Archivos",
                 bg=BG, fg=TEXT, font=("Consolas", 15, "bold")).pack(pady=(14,2))
        tk.Label(self, text="Carga 2 archivos · Únelos · Ordénalos con Insertion Sort · Guarda el resultado",
                 bg=BG, fg=MUTED, font=("Consolas", 9)).pack(pady=(0,10))

        # ── Sección de carga ──
        load_frame = tk.Frame(self, bg=CARD, padx=12, pady=10)
        load_frame.pack(fill="x", padx=16, pady=(0,8))

        for col, (label, var, cmd) in enumerate([
            ("Archivo 1", self.file1, self._cargar1),
            ("Archivo 2", self.file2, self._cargar2),
        ]):
            f = tk.Frame(load_frame, bg=CARD)
            f.grid(row=0, column=col, sticky="ew", padx=8)
            load_frame.columnconfigure(col, weight=1)

            tk.Label(f, text=label, bg=CARD, fg=MUTED,
                     font=("Consolas", 9)).pack(anchor="w")
            row = tk.Frame(f, bg=CARD)
            row.pack(fill="x")
            tk.Label(row, textvariable=var, bg=PANEL, fg=TEXT,
                     font=("Consolas", 9), anchor="w", padx=6,
                     relief="flat", width=32).pack(side="left", fill="x", expand=True)
            tk.Button(row, text="📂 Abrir", command=cmd,
                      bg=BLUE, fg="#fff", relief="flat",
                      font=("Consolas", 9), cursor="hand2",
                      activebackground="#3a7be0", padx=8, pady=3).pack(side="left", padx=4)

        # Botón generar archivos de prueba
        tk.Button(load_frame, text="🎲 Generar archivos de prueba",
                  command=self._generar_prueba,
                  bg=PANEL, fg=MUTED, relief="flat",
                  font=("Consolas", 9), cursor="hand2",
                  padx=8, pady=3).grid(row=1, column=0, columnspan=2,
                                        sticky="w", padx=8, pady=(8,0))

        # ── Vista previa de datos ──
        prev = tk.Frame(self, bg=CARD, padx=12, pady=8)
        prev.pack(fill="x", padx=16, pady=(0,6))
        prev.columnconfigure((0,1,2), weight=1)

        for col, (label, color, attr) in enumerate([
            ("Archivo 1", BLUE,  "lbl_d1"),
            ("Archivo 2", PINK,  "lbl_d2"),
            ("Unión",     GOLD,  "lbl_mg"),
        ]):
            f = tk.Frame(prev, bg=CARD)
            f.grid(row=0, column=col, sticky="ew", padx=6)
            tk.Frame(f, bg=color, height=3).pack(fill="x")
            tk.Label(f, text=label, bg=CARD, fg=color,
                     font=("Consolas", 9, "bold")).pack(anchor="w", pady=2)
            lbl = tk.Label(f, text="—", bg=PANEL, fg=TEXT,
                            font=("Consolas", 8), anchor="w",
                            wraplength=230, justify="left", padx=4)
            lbl.pack(fill="x")
            setattr(self, attr, lbl)

        # ── Canvas de visualización ──
        tk.Label(self, text="Visualización del ordenamiento",
                 bg=BG, fg=MUTED, font=("Consolas", 9)).pack(anchor="w", padx=16)

        self.canvas = BarCanvas(self, width=820, height=190)
        self.canvas.pack(fill="x", padx=16, pady=(2,0))

        # Mensaje de estado
        self.msg = tk.Label(self, text="Carga 2 archivos para comenzar.",
                             bg=BG, fg=MUTED, font=("Consolas", 9), anchor="w")
        self.msg.pack(fill="x", padx=16, pady=2)

        # Progreso
        self.progress = ttk.Progressbar(self, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", padx=16, pady=(0,4))

        # ── Controles ──
        ctrl = tk.Frame(self, bg=BG)
        ctrl.pack(fill="x", padx=16, pady=4)

        btn = dict(bg=PANEL, fg=TEXT, relief="flat", font=("Consolas", 10),
                   cursor="hand2", activebackground="#252836",
                   activeforeground=TEXT, padx=10, pady=5)

        self.btn_sort  = tk.Button(ctrl, text="▶ Ordenar", command=self._iniciar, **btn)
        self.btn_sort.pack(side="left", padx=2)
        self.btn_sort.config(bg=BLUE, fg="#fff", activebackground="#3a7be0")

        self.btn_pause = tk.Button(ctrl, text="⏸ Pausa", command=self._pausar,
                                    state="disabled", **btn)
        self.btn_pause.pack(side="left", padx=2)

        self.btn_step  = tk.Button(ctrl, text="→ Paso", command=self._paso, **btn)
        self.btn_step.pack(side="left", padx=2)

        self.btn_reset = tk.Button(ctrl, text="↺ Reset", command=self._reset, **btn)
        self.btn_reset.pack(side="left", padx=2)

        tk.Label(ctrl, text="Velocidad:", bg=BG, fg=MUTED,
                 font=("Consolas", 9)).pack(side="left", padx=(14,2))
        self.spd_var = tk.IntVar(value=6)
        ttk.Scale(ctrl, from_=1, to=10, variable=self.spd_var,
                  orient="horizontal", length=80,
                  command=lambda v: self._upd_speed()).pack(side="left")

        # ── Botón guardar ──
        self.btn_save = tk.Button(self, text="💾  Guardar archivo ordenado",
                                   command=self._guardar,
                                   bg=GREEN, fg="#fff", relief="flat",
                                   font=("Consolas", 10, "bold"), cursor="hand2",
                                   activebackground="#2eb87a", pady=7,
                                   state="disabled")
        self.btn_save.pack(fill="x", padx=16, pady=(6,2))

        # ── Leyenda ──
        leg = tk.Frame(self, bg=BG)
        leg.pack(fill="x", padx=16, pady=(4,12))
        for color, label in [(BAR_DEF,"Normal"),(BAR_KEY,"Clave"),
                              (BAR_CMP,"Comparando"),(PINK,"Moviendo"),(BAR_DONE,"Ordenado")]:
            tk.Frame(leg, bg=color, width=12, height=12).pack(side="left", padx=(6,2))
            tk.Label(leg, text=label, bg=BG, fg=MUTED,
                     font=("Consolas", 8)).pack(side="left", padx=(0,6))

    # ── Helpers ──────────────────────────────────────────
    def _upd_speed(self):
        self.speed = int(550 - self.spd_var.get() * 50)

    def _fmt(self, lst):
        s = ", ".join(str(int(x) if x == int(x) else x) for x in lst[:20])
        return s + (f"  … (+{len(lst)-20} más)" if len(lst) > 20 else "")

    # ── Carga de archivos ─────────────────────────────────
    def _cargar(self, num):
        path = filedialog.askopenfilename(
            title=f"Selecciona Archivo {num}",
            filetypes=[
                ("Todos los soportados", "*.txt *.json *.csv *.xlsx"),
                ("Texto plano",          "*.txt"),
                ("JSON",                 "*.json"),
                ("CSV",                  "*.csv"),
                ("Excel",                "*.xlsx"),
            ])
        if not path:
            return
        try:
            datos = leer_archivo(path)
            if num == 1:
                self.data1 = datos
                self.file1.set(os.path.basename(path))
                self.lbl_d1.config(text=self._fmt(datos))
            else:
                self.data2 = datos
                self.file2.set(os.path.basename(path))
                self.lbl_d2.config(text=self._fmt(datos))
            # Mostrar info de hojas si es xlsx
            ext = os.path.splitext(path)[1].lower()
            if ext == ".xlsx" and path in _xlsx_info:
                info = _xlsx_info[path]
                if num == 1:
                    self.lbl_d1.config(text=f"[Hojas: {info}]\n{self._fmt(datos)}")
                else:
                    self.lbl_d2.config(text=f"[Hojas: {info}]\n{self._fmt(datos)}")
            self._actualizar_union()
        except Exception as e:
            messagebox.showerror("Error al leer archivo", str(e))

    def _cargar1(self): self._cargar(1)
    def _cargar2(self): self._cargar(2)

    def _actualizar_union(self):
        if self.data1 or self.data2:
            self.merged = self.data1 + self.data2
            self.lbl_mg.config(text=self._fmt(self.merged))
            self.canvas.set_data(self.merged)
            self.msg.config(text=f"Unión: {len(self.merged)} elementos listos para ordenar.")

    # ── Generar archivos de prueba ────────────────────────
    def _generar_prueba(self):
        d = filedialog.askdirectory(title="Elige carpeta para guardar archivos de prueba")
        if not d:
            return
        import csv as _csv
        import openpyxl as _xl
        nums = [random.sample(range(1, 300), 15) for _ in range(4)]
        # .txt
        p1 = os.path.join(d, "prueba1.txt")
        with open(p1, "w") as f:
            f.write(", ".join(map(str, nums[0])))
        # .json
        p2 = os.path.join(d, "prueba2.json")
        with open(p2, "w") as f:
            json.dump(nums[1], f)
        # .csv con encabezado
        p3 = os.path.join(d, "prueba3.csv")
        with open(p3, "w", newline="") as f:
            w = _csv.writer(f)
            w.writerow(["numero"])
            for n in nums[2]:
                w.writerow([n])
        # .xlsx con 3 hojas (para demostrar lectura multi-hoja)
        p4 = os.path.join(d, "prueba4.xlsx")
        wb = _xl.Workbook()
        for idx_h, nombre_h in enumerate(["Hoja1", "Hoja2", "Hoja3"]):
            if idx_h == 0:
                ws = wb.active
                ws.title = nombre_h
            else:
                ws = wb.create_sheet(nombre_h)
            ws.append(["numero"])
            extra = random.sample(range(1, 300), 5)
            for n in extra:
                ws.append([n])
        wb.save(p4)
        messagebox.showinfo("Archivos creados",
            f"Se crearon 4 archivos en:\n{d}\n\n"
            "  prueba1.txt\n  prueba2.json\n  prueba3.csv\n  prueba4.xlsx\n\n"
            "Cárgalos con los botones 📂")

    # ── Ordenamiento ─────────────────────────────────────
    def _iniciar(self):
        if not self.merged:
            messagebox.showwarning("Sin datos", "Carga al menos un archivo primero.")
            return
        if not self.steps:
            self.steps = insertion_sort_steps(self.merged)
            self.idx   = 0
            self.progress["maximum"] = len(self.steps)
        self.running = True
        self.btn_pause.config(state="normal")
        self.btn_sort.config(state="disabled")
        self._auto()

    def _auto(self):
        if not self.running:
            return
        if self.idx < len(self.steps):
            self._render()
            self.idx += 1
            self._after = self.after(self.speed, self._auto)
        else:
            self.running = False
            self.btn_save.config(state="normal")
            self.btn_pause.config(state="disabled")

    def _render(self):
        if self.idx >= len(self.steps):
            return
        kind, pos, arr, msg = self.steps[self.idx]
        hl = {}
        if kind == "key":
            hl = {pos: BAR_KEY}
        elif kind == "cmp":
            hl = {pos: BAR_CMP}
        elif kind in ("move", "place"):
            hl = {pos: PINK}
        elif kind == "done":
            hl = {i: BAR_DONE for i in range(len(arr))}
        self.canvas.set_data(arr, hl)
        self.msg.config(text=msg)
        self.progress["value"] = self.idx + 1

    def _pausar(self):
        self.running = False
        if self._after:
            self.after_cancel(self._after)
        self.btn_sort.config(state="normal", text="▶ Continuar")
        self.btn_pause.config(state="disabled")

    def _paso(self):
        if not self.merged:
            return
        if not self.steps:
            self.steps = insertion_sort_steps(self.merged)
            self.idx   = 0
            self.progress["maximum"] = len(self.steps)
        if self.idx < len(self.steps):
            self._render()
            self.idx += 1

    def _reset(self):
        self.running = False
        if self._after:
            self.after_cancel(self._after)
        self.steps = []
        self.idx   = 0
        self.progress["value"] = 0
        self.btn_sort.config(state="normal", text="▶ Ordenar")
        self.btn_pause.config(state="disabled")
        self.btn_save.config(state="disabled")
        self._actualizar_union()
        self.msg.config(text="Reset. Presiona ▶ Ordenar para volver a empezar.")

    # ── Guardar resultado ─────────────────────────────────
    def _guardar(self):
        if not self.steps:
            return
        # El arreglo final está en el último paso
        _, _, resultado, _ = self.steps[-1]
        path = filedialog.asksaveasfilename(
            title="Guardar archivo ordenado",
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt"), ("JSON", "*.json")])
        if not path:
            return
        try:
            guardar_archivo(path, resultado)
            messagebox.showinfo("Guardado",
                                 f"Archivo guardado exitosamente:\n{path}\n\n"
                                 f"{len(resultado)} elementos ordenados.")
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))


# ── Entry point ───────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()